"""
Script para crear un archivo Excel de plantilla para la carga masiva de datos.
Este archivo puede usarse como ejemplo para cargar datos tributarios.
"""

import pandas as pd
from datetime import datetime

try:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[ADVERTENCIA] openpyxl no esta disponible. El archivo se creara sin formato avanzado.")

def crear_plantilla_excel():
    """Crea un archivo Excel de plantilla con el formato correcto"""
    
   
    datos_ejemplo = {
        'Nombre': [
            'Inversión Renta Fija Banco Estado',
            'Fondo Mutuo Renta Variable',
            'Depósito a Plazo',
            'Acciones Empresa ABC',
            'Bonos Corporativos',
            'ETF Internacional',
            'Fondo de Inversión Local',
            'Certificado de Depósito',
            'Letra del Tesoro',
            'Obligación Bancaria'
        ],
        'Monto': [
            1000000.50,
            2500000.75,
            500000.00,
            1500000.25,
            800000.50,
            3000000.00,
            1200000.00,
            600000.75,
            900000.00,
            1750000.50
        ],
        'Factor': [
            1.05,
            1.15,
            1.02,
            1.25,
            1.08,
            1.20,
            1.10,
            1.03,
            1.01,
            1.12
        ],
        'Fecha': [
            '2024-01-15',
            '2024-02-20',
            '2024-03-10',
            '2024-04-05',
            '2024-05-12',
            '2024-06-18',
            '2024-07-22',
            '2024-08-30',
            '2024-09-15',
            '2024-10-25'
        ]
    }
    
   
    df = pd.DataFrame(datos_ejemplo)
    
    
    nombre_archivo = 'plantilla_carga_datos.xlsx'
    
   
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')
        
       
        worksheet = writer.sheets['Datos']
        
       
        if OPENPYXL_AVAILABLE:
            try:
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
            
                for idx, col in enumerate(df.columns, 1):
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(col)
                    ) + 3
                    col_letter = get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].width = min(max_length, 50)
                
                
                for row in range(2, len(df) + 2):
                   
                    monto_cell = worksheet.cell(row=row, column=2)
                    monto_cell.number_format = '#,##0.00'
                    
                   
                    factor_cell = worksheet.cell(row=row, column=3)
                    factor_cell.number_format = '0.00'
                    
                 
                    fecha_cell = worksheet.cell(row=row, column=4)
                    fecha_cell.number_format = 'YYYY-MM-DD'
                
               
                instrucciones = writer.book.create_sheet("Instrucciones")
                
                instrucciones_texto = [
                    ["INSTRUCCIONES PARA CARGA MASIVA DE DATOS"],
                    [""],
                    ["COLUMNAS REQUERIDAS:"],
                    ["- Nombre: Obligatorio. Nombre o descripción del dato tributario."],
                    ["- Monto: Opcional. Monto numérico (ejemplo: 1000000.50)"],
                    ["- Factor: Opcional. Factor numérico (ejemplo: 1.05)"],
                    ["- Fecha: Opcional. Fecha en formato YYYY-MM-DD (ejemplo: 2024-01-15)"],
                    [""],
                    ["FORMATO DE ARCHIVO:"],
                    ["- El archivo debe ser .xlsx o .csv"],
                    ["- La primera fila debe contener los nombres de las columnas"],
                    ["- Los nombres de columnas pueden estar en español o inglés"],
                    ["- Tamaño máximo del archivo: 10MB"],
                    [""],
                    ["VARIACIONES DE NOMBRES DE COLUMNAS ACEPTADAS:"],
                    ["- Nombre: Nombre, Name, Nombre Dato, Descripción, Desc, Dato, Item"],
                    ["- Monto: Monto, Amount, Valor, Value, Precio, Price, Importe"],
                    ["- Factor: Factor, Factor_, Multiplicador, Multiplier, Ratio, Coeficiente"],
                    ["- Fecha: Fecha, Date, Fecha Dato, Fecha Creación, Created At"],
                    [""],
                    ["NOTAS:"],
                    ["- La columna 'Nombre' es obligatoria"],
                    ["- Las demás columnas son opcionales"],
                    ["- El sistema detectará automáticamente las columnas"],
                    ["- Puedes eliminar las filas de ejemplo y agregar tus propios datos"],
                    ["- Las fechas pueden estar en cualquier formato estándar"],
                    [""],
                    ["MODOS DE CARGA:"],
                    ["- Crear: Crea nuevos registros"],
                    ["- Actualizar: Actualiza registros existentes por nombre"],
                ]
                
                for row_num, row_data in enumerate(instrucciones_texto, 1):
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = instrucciones.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        
                        
                        if row_num == 1:
                            cell.font = Font(bold=True, size=14, color="366092")
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif row_num in [3, 9, 15, 20, 26]:
                            cell.font = Font(bold=True, size=11)
                            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                
                instrucciones.column_dimensions['A'].width = 80
            except Exception as e:
                print(f"[ADVERTENCIA] No se pudo aplicar formato avanzado: {e}")
                print("   El archivo se creara con formato basico.")
        
    print(f"[OK] Archivo Excel creado exitosamente: {nombre_archivo}")
    print(f"[INFO] Ubicacion: {nombre_archivo}")
    print(f"[INFO] Filas de ejemplo: {len(df)}")
    print(f"[INFO] Columnas: {', '.join(df.columns.tolist())}")
    print("\n[INFO] Puedes usar este archivo para:")
    print("   - Ver el formato correcto")
    print("   - Agregar tus propios datos")
    print("   - Cargarlo en el sistema")

if __name__ == "__main__":
    try:
        crear_plantilla_excel()
    except Exception as e:
        print(f"[ERROR] Error al crear el archivo Excel: {e}")
        import traceback
        traceback.print_exc()

